# just for test
import openpyxl
from openpyxl.styles import Font
from bs4 import BeautifulSoup
import requests

class Stock:
    #建構式
    def __init__(self, *stock_numbers):
        self.stock_numbers = stock_numbers
    
    #爬取
    def scrape(self):
	
        result = list()  #最終結果
 
        for stock_number in self.stock_numbers:
            
            response = requests.get("https://tw.stock.yahoo.com/q/q?s=" + stock_number)
            soup = BeautifulSoup(response.text.replace("加到投資組合", ""), "lxml")
                
            stock_date = soup.find("font", {"class": "tt"}).getText().strip()[-9:]  #資料日期
                
            tables = soup.find_all("table")[2]  #取得網頁中第三個表格(索引從0開始計算)
            tds = tables.find_all("td")[0:11]  #取得表格中1到10格
            
            result.append((stock_date,) + tuple(td.getText().strip() for td in tds))                
        
        return result

    def export(self, stocks):
        wb = openpyxl.Workbook()
        sheet = wb.create_sheet("Yahoo股市", 0)
    
        response = requests.get(
            "https://tw.stock.yahoo.com/q/q?s=2451")
        soup = BeautifulSoup(response.text, "lxml")
    
        tables = soup.find_all("table")[2]
        ths = tables.find_all("th")[0:11]
        titles = ("資料日期",) + tuple(th.getText() for th in ths)
        sheet.append(titles)
    
        for index, stock in enumerate(stocks):
            sheet.append(stock)
    
            if "△" in stock[6]:
                sheet.cell(row=index+2, column=7).font = Font(color='FF0000')  #儲存格字體顯示紅色
            elif "▽" in stock[6]:
                sheet.cell(row=index+2, column=7).font = Font(color='00A600')  #儲存格字體顯示綠色
    
        wb.save("yahoostock.xlsx")

stock = Stock('2451', '2454', '2369')  # 建立Stock物件
stock.export(stock.scrape())  #將爬取的結果匯出成Excel檔案